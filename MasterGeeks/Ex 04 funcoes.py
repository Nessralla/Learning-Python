# Programa que controla a coleção de naves de STAR WARS

# Passo a passo do código:
# 1: Chamamos a função para criar o dicionário com base na planilha (ler_plan_mudar_dict)
# 2: Em seguido iniciamos com o menu principal(menu_controle)
# 2.1: Cadastrar
# 2.2: Visualizar
# 2.3: Trocar naves
# 2.4: Encerrar

relNaves = {
    'U-Wing': ['Caça Estelar','19.5x9.5x35','Wing Miniature Game', 'Usado', 2],
    'TIE Fighter': ['Caça Estelar','16x3x10','Wing Miniature Game', 'Novo', 2], 
    'C-ROC': ['Cruzador Leve', '23.5x41.8x14', 'X-Wing Miniature Game', 'Novo', 1], 
    'TIE Punisher': ['Caça Esterlar', '19x7x10', 'X-Wing Miniature Game', 'Usado', 3], 
    'X-Wing': ['Caça Esterlar', '19x15x10', 'Crystal Living', 'Novo', 0], 
    'Slave 1': ['Nave de Patrulha', '20x12x8', 'X-Wing Miniature Game', 'Novo', 0], 
    'TIE Silencer': ['Caça Esterlar', '18x9.5x7', 'X-Wing Miniature Game', 'Novo', 0], 
    'YT-1300 Millenium Falcon': ['Cargueiro Leve', '45x30x20', 'X-Wing Miniature Game', 'Usado', 0]
    }

from time import sleep

import xlsxwriter
import openpyxl



def separadores():
    """
        Função simples para criar separadores de texto para ficar mais bonito
    """
    print('-='*40)


def mostrarColeção():
    """
        Função que exibe a coleção de naves guardada em um dicionário
    """
    separadores()
    print(" "*12,"Coleção de Naves do Sr. Borges"," "*12)
    separadores()
    print("Nome","Tipo","Tamanho(cm)(CxLxA)","Coleção","Estado","Qtd")
    for i in relNaves:
        print(f"\n{i}", end='      ')
        for j in range(len(relNaves["U-Wing"])):
            print(f"{relNaves[i][j]}",end="  ")

def menu_controle():
    """
        Função que roda o programa principal, para que o usuário escolha o que deseja fazer
    """

    while True:
        print("Escolha uma das 4 opções abaixo:")
        sleep(1)
        print("""
            1 : Cadastrar nova nave
            2 : Visualizar as naves cadastradas
            3 : Trocar uma nave por outra (neste caso, a nova nave deve ser cadastrada antecipadamente, com 
            a quantidade 0!)
            4 : Encerrar o programa
        """)
        opc = input("\n Opção: ")
        if opc == "1":
            
            print("Para cadastrar uma nova nave, siga os passos abaixo: ")
            sleep(0.5)
            cadastrar_nave()
            
            
        elif opc == "2":
            print("Exibindo a base de naves")
            sleep(0.5)
            mostrarColeção()
            
            
        elif opc == "3":
            # Chama a função de trocar a nave
            print("Chamando a função para trocar naves")
            sleep(0.5)
            trocar_nave()
        elif opc == "4":
            
            print("Encerrando o programa! ")
            sleep(0.5)
            escrever_bancoDados(r"C:\Users\Usuario\MyApp\Learning Python\planilha.xlsx")
            sleep(0.5)
            print("Até logo")
            sleep(0.5)
            break
        
def cadastrar_nave():
    """
        Função para cadastrar uma nova nave no dicionário
    """
    print("Iniciando o cadastramento de uma nave!")
    sleep(0.5)
    conf = "N"
    
    name = input("\nDigite o nome da nave: ")

    if confirmacao(name,"Digite novamente a nave: "):
        sleep(0.5)
        relNaves[name] = list()
        opc = (input(f"Digite qual o tipo da nave {name}: "))
        if confirmacao(opc,"Digite o tipo da nave: "):
            relNaves[name].append(opc)

        opc = input(f"Digite o tamanho da {name}, em cm, no formato COMPxLARGxALT: ")
        if confirmacao(opc,"Digite o tamanho da nave, em cm, no formato COMPxLARGxALT: "):
            relNaves[name].append(opc)

        opc = input(f"Digite qual a coleção da {name}: ")
        if confirmacao(opc,f"Digite qual a coleção da {name}: "):
            relNaves[name].append(opc)

        opc = input(f"Digite qual o Estado da {name}: ")
        if confirmacao(opc,f"Digite qual o Estado da {name}: "):
            relNaves[name].append(opc)

        opc = input(f"Digite qual a quantidade que tem da {name}: ")
        if confirmacao(opc,f"Digite qual a quantidade da {name}: "):
            relNaves[name].append(opc)

def ler_plan_mudar_dict(local):
    
    """
        Para rodar o programa com o arquivo salvo, é necessário primeiro ler o arquivo e atualizar o dicionário
        Função para ler a planilha e consequentemente alterar o dicionário, para rodar o código.
        Recebe um parâmento, que é o endereço do arquivo.
    """
    
    wb = openpyxl.load_workbook(local)
    sheet = wb["Sheet1"]
    
    relNaves.clear() # Vamos limpar o dicionário do primeiro uso, ou se por um bug tiver ficado salvo

    for row in sheet.iter_rows(max_row=sheet.max_row):
        
        relNaves[row[0].value] = list()
        
        for cell in row:
            if row[0].value == cell.value:
                None
            else:
                relNaves[row[0].value].append(cell.value)
    
    print(relNaves)


def confirmacao(var1,exp=" "):
    """
        Função para confirmação do input digitado:
        var1 = input digitado pelo usuário, anteriormente, para validação
        exp (opcional) = recomendado utilizar, serve para relembrar o usuário de qual dado ele esta inserindo
    """
    sleep(0.5)
    conf = input(f"\nVocê digitou {var1}. Deseja confirmar ? [S/N] :  ")    
    while conf.upper() != "S":
        var1 = input(exp)
        sleep(0.5)
        conf = input(f"\nVocê digitou {var1}. Deseja confirmar ? [S/N] :  ")
    return True
    

def escrever_bancoDados(local):
    """
        Função que reescreve toda a planilha com base no dicionário
    """
    row = 0
    col = 0
    file = xlsxwriter.Workbook(local)
    aba = file.add_worksheet()
    for nave in relNaves:
        aba.write(row,col,nave)
        
        for i in range(1,len(relNaves[nave])+1):
            aba.write(row,(col+i),relNaves[nave][i-1])
        
        row += 1
    
    file.close()


def trocar_nave():
    print(f"Vamos iniciar a troca das naves, primeiramente, escolha abaixo a nave que irá entregar: ")
    for nave in relNaves:
        if relNaves[nave][4] > 1:
            print(f"Nave: {nave} com a quantidade: {relNaves[nave][4]} ")
    opcA = input("Digite o nome da nave, atenção pois deve ser idêntico: ")
    if confirmacao(opcA,"Digite novamente a nave: "):
        try:
            if relNaves[opcA][4] > 1:
                conf = input(f"Da nave {opcA} será reduzida uma unidade, deseja confirmar? [S/N]  ")
                confirmacao(conf,f"Da nave {opcA} será reduzida uma unidade, deseja confirmar? [S/N]  ")
                relNaves[opcA][4] -= 1
                print(f"A nave {opcA} agora possui quantidade: {relNaves[opcA][4]} ")
                sleep(1)
                print("Agora escolha a nave que irá receber na troca, entre as opções abaixo: ")
                sleep(0.5)
                for nave in relNaves:
                    print(nave,end="  -  ")
                opcB = input("\n Digite o nome da nave, atenção pois deve ser idêntico: ")
                if confirmacao(opcB,"Digite novamente a nave: "):
                    try:
                        if opcB in relNaves.keys():
                            print(f"Será adicionada a nave {opcB} uma unidade, deseja confirmar: [S/N]")
                            confirmacao(conf,f"Será adicionada a nave {opcB} uma unidade, deseja confirmar: [S/N]")
                            relNaves[opcB][4] += 1
                            print(f"A nave {opcB} agora possui quantidade: {relNaves[opcB][4]} ")
                    except:
                        print(f"A nave {opcB} não existe na relação, a operação será desfeita. Retorne ao menu e refaça novamente")
                        relNaves[opcA][4] += 1
            else:
                print(f"A {opcA} não possui unidades extras disponíveis para trocar")

        except:
            print("Voce passou uma nave inexistente")

    
menu_controle()


